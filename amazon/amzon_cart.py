from selenium import webdriver
import selenium.webdriver.support.ui as ui
import time
import openpyxl

"fire fox driver"
# driver=webdriver.Firefox(executable_path="/usr/local/bin/geckodriver")
"chrome driver"
driver=webdriver.Chrome(executable_path="/usr/local/bin/chromedriver")

url="https://www.amazon.in/"

driver.get(url)
driver.maximize_window()
time.sleep(3)

#configure spreadsheet path
excel_path="/home/vigneshkumar/Documents/userdata.xlsx"

b=openpyxl.load_workbook(excel_path)

#open pyxl
#get active sheet
sheet=b.active
#get cell address of email within active sheet
email=sheet.cell(row=2,column=1)
#get cell address of password withiin active sheet
password=sheet.cell(row=2,column=2)
#get values
user_email=email.value
user_password=password.value


driver.find_element_by_xpath("//a[@id='nav-link-accountList']").click()
time.sleep(3)
email=driver.find_element_by_xpath("//input[@id='ap_email']")
email.send_keys(user_email)
time.sleep(3)
conti=driver.find_element_by_xpath("//input[@id='continue']")
conti.click()
time.sleep(3)
pas_ssword=driver.find_element_by_xpath("//input[@id='ap_password']")
pas_ssword.send_keys(user_password)
time.sleep(3)
driver.find_element_by_xpath("//input[@id='signInSubmit']").click()
time.sleep(3)

driver.find_element_by_xpath("//input[@id='twotabsearchtextbox']").click()
driver.find_element_by_xpath("//input[@id='twotabsearchtextbox']").send_keys("i phone")
driver.find_element_by_xpath("//input[@id='nav-search-submit-button']").click()

windows_before=driver.window_handles[0]


"scrolling for specific location"
driver.execute_script("window.scrollBy(1000,1400)","")
time.sleep(3)
select_phone=driver.find_element_by_xpath("//span[normalize-space()='New Apple iPhone 12 (128GB) - Purple']")
select_phone.click()
time.sleep(3)


windows_after=driver.window_handles[1]

driver.switch_to.window(windows_before)
time.sleep(3)

driver.switch_to.window(windows_after)

driver.implicitly_wait(5)
wait = ui.WebDriverWait(driver, 10)
wait.until(lambda driver: driver.find_element_by_id('add-to-cart-button'))
pwd = driver.find_element_by_id('add-to-cart-button')
pwd.click()